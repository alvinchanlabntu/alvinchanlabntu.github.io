from __future__ import annotations

import json
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MEMBER_SHEET = ROOT / "Lab Member Information Form (Responses).xlsx"
PUBLICATION_SHEET = ROOT / "publication_form_fields.xlsx"
PHOTO_DIR = ROOT / "Profile Photo (File responses)"
FIGURE_DIR = ROOT / "figures"
PAPER_DIR = ROOT / "papers"

MEMBER_DIR = ROOT / "_members"
MEMBER_IMAGE_DIR = ROOT / "images" / "members"
PUBLICATION_IMAGE_DIR = ROOT / "images" / "publications"
PUBLICATION_FILE_DIR = ROOT / "files" / "publications"

ROLE_MAP = {
    "Principal Investigator": "principal-investigator",
    "Postdoctoral Researcher": "postdoc",
    "PhD Student": "phd",
    "Research Assistant": "research-assistant",
}

PUBLICATION_SLUGS = [
    "strategic-multimodal-alignment",
    "designing-lipid-nanoparticles",
    "smiself-valid-molecules",
    "in-context-representation-learning",
    "automated-creativity-evaluation",
    "partial-information-decomposition",
]

PUBLICATION_DATES = [
    "2026-03-14",
    "2025-08-14",
    "2025-11-01",
    "2025-12-01",
    "2026-07-01",
    "2026-07-01",
]

PUBLICATION_IDS = [
    "doi:10.1609/aaai.v40i25.39248",
    "doi:10.1038/s41565-025-01975-4",
    "doi:10.18653/v1/2025.emnlp-main.1350",
    None,
    "doi:10.18653/v1/2026.acl-long.1061",
    None,
]

PUBLICATION_TAGS = [
    ["multimodal learning", "representation alignment", "machine learning"],
    ["nanomedicine", "lipid nanoparticles", "therapeutic discovery"],
    ["molecular generation", "language models", "cheminformatics"],
    ["multimodal AI", "in-context learning", "molecular AI"],
    ["creativity", "language models", "evaluation"],
    ["multimodal AI", "interpretability", "information theory"],
]

MEMBER_ALIASES = {
    "Zhang Tianle": ["Tianle Zhang"],
    "Fang Wanlong": ["Wanlong Fang"],
    "Tao Wen": ["Wen Tao"],
    'Weerin "Mata" Ngochanthra': ["Weerin Ngochanthra"],
}

LAB_PUBLICATION_AUTHORS = {
    "Alvin Chan",
    "Tianle Zhang",
    "Wanlong Fang",
    "Wen Tao",
}


def yaml_scalar(value: object) -> str:
    return json.dumps(value, ensure_ascii=False)


def slugify(value: str) -> str:
    value = value.lower().replace('"', "")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def find_photo(name: str) -> Path:
    matches = [path for path in PHOTO_DIR.iterdir() if path.is_file() and path.stem == name]
    if len(matches) != 1:
        raise RuntimeError(f"Expected one photo for {name!r}, found {matches}")
    return matches[0]


def copy_member_photo(name: str, slug: str) -> str:
    source = find_photo(name)
    if source.suffix.lower() == ".heic":
        destination = MEMBER_IMAGE_DIR / f"{slug}.jpg"
        with tempfile.TemporaryDirectory() as temp_dir:
            subprocess.run(
                ["qlmanage", "-t", "-s", "1600", "-o", temp_dir, str(source)],
                check=True,
                stdout=subprocess.DEVNULL,
            )
            rendered = Path(temp_dir) / f"{source.name}.png"
            subprocess.run(
                [
                    "sips",
                    "-s",
                    "format",
                    "jpeg",
                    "-s",
                    "formatOptions",
                    "85",
                    str(rendered),
                    "--out",
                    str(destination),
                ],
                check=True,
                stdout=subprocess.DEVNULL,
            )
    else:
        destination = MEMBER_IMAGE_DIR / f"{slug}{source.suffix.lower()}"
        shutil.copy2(source, destination)
    return destination.relative_to(ROOT).as_posix()


def parse_homepage(value: str) -> tuple[str, str] | None:
    value = value.strip()
    match = re.search(r"https?://\S+", value)
    if not match:
        return None
    url = match.group(0).rstrip(".,)")
    scholar = re.search(r"scholar\.google\.[^/]+/citations\?[^#]*user=([^&#]+)", url, flags=re.IGNORECASE)
    if scholar:
        return "google-scholar", scholar.group(1)
    linkedin = re.search(r"linkedin\.com/in/([^/?#]+)", url, flags=re.IGNORECASE)
    if linkedin:
        return "linkedin", linkedin.group(1).rstrip("/")
    return "home-page", url


def member_markdown(row: dict[str, object]) -> tuple[str, str]:
    name = str(row["Full Name"]).strip()
    position = str(row["Position"]).strip()
    team = str(row["Team"]).strip()
    email = str(row["Email Address"]).strip()
    fun_fact = str(row["Fun Fact"]).strip()
    homepage = str(row["Homepage"] or "").strip()
    slug = slugify(name)
    image = copy_member_photo(name, slug)
    role = ROLE_MAP[position]

    front_matter = [
        "---",
        f"name: {yaml_scalar(name)}",
        f"image: {yaml_scalar(image)}",
        f"role: {yaml_scalar(role)}",
        f"description: {yaml_scalar(position if team == 'Not applicable' else f'{position} | {team}')}",
        f"team: {yaml_scalar(team)}",
        f"affiliation: {yaml_scalar('Nanyang Technological University')}",
    ]

    aliases = MEMBER_ALIASES.get(name, [])
    if aliases:
        front_matter.append("aliases:")
        front_matter.extend(f"  - {yaml_scalar(alias)}" for alias in aliases)

    front_matter.extend(["links:", f"  email: {yaml_scalar(email)}"])
    parsed_homepage = parse_homepage(homepage)
    if parsed_homepage:
        kind, link = parsed_homepage
        front_matter.append(f"  {kind}: {yaml_scalar(link)}")
    front_matter.append("---")

    if name == "Alvin Chan":
        body = (
            "Alvin Chan is an Assistant Professor jointly appointed at the College of "
            "Computing and Data Science and the Lee Kong Chian School of Medicine at NTU. "
            "His research develops generative and multimodal AI for therapeutic discovery, "
            "precision medicine, nanomedicine, and RNA therapeutics."
        )
    else:
        team_phrase = "the lab" if team == "Not applicable" else f"the lab's {team}"
        body = f"{name} is a {position} with {team_phrase}."

    if fun_fact:
        body += f"\n\n**Outside the lab:** {fun_fact}"

    return slug, "\n".join(front_matter) + f"\n\n{body}\n"


def first_sentences(text: str, limit: int = 420) -> str:
    if len(text) <= limit:
        return text
    candidate = text[:limit]
    sentence_end = max(candidate.rfind(". "), candidate.rfind("? "), candidate.rfind("! "))
    if sentence_end > 180:
        return candidate[: sentence_end + 1]
    return candidate.rstrip() + "..."


def yaml_mapping(entry: dict[str, object]) -> list[str]:
    lines: list[str] = []
    for key, value in entry.items():
        if value is None:
            continue
        if isinstance(value, list):
            lines.append(f"  {key}:")
            for item in value:
                if isinstance(item, dict):
                    lines.append("    -")
                    for child_key, child_value in item.items():
                        lines.append(f"      {child_key}: {yaml_scalar(child_value)}")
                else:
                    lines.append(f"    - {yaml_scalar(item)}")
        else:
            lines.append(f"  {key}: {yaml_scalar(value)}")
    return lines


def publication_entries() -> list[dict[str, object]]:
    workbook = load_workbook(PUBLICATION_SHEET, read_only=True, data_only=True)
    sheet = workbook["Form Fields"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    entries: list[dict[str, object]] = []

    for index, values in enumerate(rows[1:], start=1):
        row = dict(zip(headers, values, strict=True))
        slug = PUBLICATION_SLUGS[index - 1]
        figure = next(FIGURE_DIR.glob(f"{index:02d}_*.png"))
        paper = next(PAPER_DIR.glob(f"{index:02d}_*.pdf"))
        image_destination = PUBLICATION_IMAGE_DIR / f"{slug}.png"
        paper_destination = PUBLICATION_FILE_DIR / f"{slug}.pdf"
        shutil.copy2(figure, image_destination)
        shutil.copy2(paper, paper_destination)

        authors = [part.strip() for part in str(row["Authors"]).split(",")]
        authors = [f"**{author}**" if author in LAB_PUBLICATION_AUTHORS else author for author in authors]
        abstract = str(row["Abstract"]).strip()
        url = str(row["URL"]).strip()

        entries.append(
            {
                "id": PUBLICATION_IDS[index - 1],
                "title": str(row["Title"]).strip(),
                "authors": authors,
                "publisher": str(row["Venue"]).strip(),
                "date": PUBLICATION_DATES[index - 1],
                "link": url,
                "type": "paper",
                "description": first_sentences(abstract),
                "abstract": abstract,
                "image": image_destination.relative_to(ROOT).as_posix(),
                "buttons": [
                    {"type": "website", "text": "Publication", "link": url},
                    {
                        "type": "paper",
                        "text": "PDF",
                        "link": "/" + paper_destination.relative_to(ROOT).as_posix(),
                    },
                ],
                "tags": PUBLICATION_TAGS[index - 1],
            }
        )

    return sorted(entries, key=lambda entry: str(entry["date"]), reverse=True)


def write_publication_yaml(entries: list[dict[str, object]]) -> None:
    body = []
    for entry in entries:
        body.append("-")
        body.extend(yaml_mapping(entry))
    content = "\n".join(body) + "\n"
    (ROOT / "_data" / "sources.yaml").write_text(content, encoding="utf-8")
    (ROOT / "_data" / "citations.yaml").write_text(
        "# DO NOT EDIT, GENERATED FROM _data/sources.yaml\n\n" + content,
        encoding="utf-8",
    )


def main() -> None:
    for directory in [MEMBER_DIR, MEMBER_IMAGE_DIR, PUBLICATION_IMAGE_DIR, PUBLICATION_FILE_DIR]:
        directory.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(MEMBER_SHEET, read_only=True, data_only=True)
    sheet = workbook["Form responses 1"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]

    for values in rows[1:]:
        row = dict(zip(headers, values, strict=True))
        slug, content = member_markdown(row)
        (MEMBER_DIR / f"{slug}.md").write_text(content, encoding="utf-8")

    entries = publication_entries()
    write_publication_yaml(entries)
    print(f"Generated {len(rows) - 1} members and {len(entries)} publications")


if __name__ == "__main__":
    main()
